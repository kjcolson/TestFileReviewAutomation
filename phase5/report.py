"""
phase5/report.py

Console display, Excel report, and JSON manifest for Phase 5.
ASCII-only output (no Unicode box-drawing characters), matching Phase 3/4 convention.
"""

from __future__ import annotations

import json
from datetime import date
from pathlib import Path
from typing import Any


_BOX_WIDTH = 69
_SEV_ORDER = {"CRITICAL": 0, "HIGH": 1, "MEDIUM": 2, "LOW": 3, "INFO": 4}

_PHASE4_CHECK_LABELS = {
    "C0": "Trans <-> Charges",
    "C0a": "Charge ID Linkage",
    "C0b": "Payment Balance",
    "C1": "Billing <-> GL",
    "C2": "Billing <-> Payroll",
    "C3": "Billing <-> Scheduling",
    "C3a": "Location/Dept Cross-Ref",
    "C3b": "Provider NPI Cross-Ref",
    "C3c": "Patient ID Cross-Ref",
    "C4": "Payroll <-> GL",
    "C5": "Scheduling <-> GL",
}


# ---------------------------------------------------------------------------
# Console formatting helpers
# ---------------------------------------------------------------------------

def _sev_icon(sev: str) -> str:
    return {"CRITICAL": "X", "HIGH": "!", "MEDIUM": "o", "LOW": ".", "INFO": "."}.get(sev, ".")


def _hr(char: str = "-") -> str:
    return "+" + char * (_BOX_WIDTH - 2) + "+"


def _row(text: str) -> str:
    return "| " + text.ljust(_BOX_WIDTH - 4) + " |"


def _kv_row(key: str, val: str, key_width: int = 22) -> str:
    k = key.ljust(key_width)
    v = str(val)[:_BOX_WIDTH - key_width - 7].ljust(_BOX_WIDTH - key_width - 7)
    return f"| {k}| {v}|"


def _truncate(s: str, max_len: int = _BOX_WIDTH - 12) -> str:
    return s[:max_len] + "..." if len(s) > max_len else s


# ---------------------------------------------------------------------------
# Console: Executive Summary
# ---------------------------------------------------------------------------

def _render_executive_summary(unified: dict, readiness_result: dict) -> str:
    lines = []
    lines.append(_hr())
    lines.append(_row("EXECUTIVE SUMMARY"))
    lines.append(_hr())
    lines.append(_kv_row("Client", unified.get("client", "")))
    lines.append(_kv_row("Round", unified.get("round", "")))
    lines.append(_kv_row("Test Month", unified.get("test_month", "")))
    bf = unified.get("billing_format", "unknown")
    lines.append(_kv_row("Billing Format", bf.title() if bf else "Unknown"))
    aligned = "Yes" if unified.get("month_aligned", True) else "No"
    lines.append(_kv_row("Month Aligned", aligned))
    lines.append(_hr())

    overall = readiness_result.get("overall", "Unknown")
    lines.append(_row(f"READINESS: {overall.upper()}"))
    reason = readiness_result.get("reason", "")
    if reason:
        lines.append(_row(f"  {_truncate(reason, _BOX_WIDTH - 8)}"))
    lines.append(_hr())

    # Source summary table
    hdr = f"| {'Source':<21}| {'Status':<12}| {'CRIT':>5}| {'HIGH':>5}| {' MED':>5}| {'Total':>6} |"
    lines.append(hdr)
    lines.append(_hr())

    per_source = readiness_result.get("per_source", {})
    # Display order
    display_order = ["billing", "scheduling", "payroll", "gl", "quality", "patient_satisfaction", "cross_source"]
    missing = readiness_result.get("missing_sources", [])

    for group in display_order:
        if group in per_source:
            ps = per_source[group]
            name = ps.get("display_name", group.title())[:20]
            status = ps.get("status", "N/A")[:11]
            crit = ps.get("critical", 0)
            high = ps.get("high", 0)
            med = ps.get("medium", 0)
            total = ps.get("total", 0)
            lines.append(f"| {name:<21}| {status:<12}| {crit:>5}| {high:>5}| {med:>5}| {total:>6} |")
        elif group in missing:
            from phase5.missing_sources import _SOURCE_GROUP
            name = group.replace("_", " ").title()[:20]
            lines.append(f"| {name:<21}| {'NOT SUBMITTED':<12}| {'  -':>5}| {'  -':>5}| {'  -':>5}| {'  -':>6} |")

    lines.append(_hr())

    # Totals
    tc = readiness_result.get("total_counts", {})
    lines.append(f"| {'TOTAL':<21}| {'':12}| {tc.get('CRITICAL', 0):>5}| {tc.get('HIGH', 0):>5}| {tc.get('MEDIUM', 0):>5}| {sum(v for k, v in tc.items() if k != 'INFO'):>6} |")
    lines.append(_hr())

    # Date range per source
    sources = unified.get("sources", {})
    has_dates = any(
        sources.get(g, {}).get("date_range", {}).get("min")
        for g in display_order if g != "cross_source"
    )
    if has_dates:
        lines.append("")
        lines.append(_hr())
        lines.append(_row("DATE RANGES"))
        lines.append(_hr())
        for group in display_order:
            if group == "cross_source":
                continue
            sdata = sources.get(group, {})
            dr = sdata.get("date_range", {})
            if not dr.get("min"):
                continue
            name = sdata.get("display_name", group.title())[:20]
            col = dr.get("date_column", "")[:20]
            date_str = f"{dr['min']} to {dr['max']}"
            lines.append(_row(f"  {name:<16} {col:<20} {date_str}"))
            if dr.get("note"):
                lines.append(_row(f"    NOTE: {_truncate(dr['note'], _BOX_WIDTH - 14)}"))
        lines.append(_hr())

    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Console: Client Issue List
# ---------------------------------------------------------------------------

def _render_issue_list(client_issues: list[dict], max_display: int = 20) -> str:
    lines = []
    total = len(client_issues)
    showing = min(total, max_display)
    lines.append(_hr())
    lines.append(_row(f"CLIENT ISSUE LIST ({showing} of {total} issues)"))
    lines.append(_hr())

    for issue in client_issues[:max_display]:
        sev = issue.get("severity", "INFO")
        icon = _sev_icon(sev)
        line = issue.get("line", "")
        # Wrap long lines
        if len(line) > _BOX_WIDTH - 8:
            first = line[:_BOX_WIDTH - 8]
            rest = line[_BOX_WIDTH - 8:]
            lines.append(_row(f"  {icon} {first}"))
            while rest:
                chunk = rest[:_BOX_WIDTH - 10]
                rest = rest[_BOX_WIDTH - 10:]
                lines.append(_row(f"      {chunk}"))
        else:
            lines.append(_row(f"  {icon} {line}"))

    if total > max_display:
        lines.append(_row(f"  ... and {total - max_display} more issues"))

    lines.append(_hr())
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Console: Resubmission Checklist
# ---------------------------------------------------------------------------

def _render_checklist(checklist_items: list[dict]) -> str:
    lines = []
    lines.append(_hr())
    lines.append(_row("RESUBMISSION CHECKLIST"))
    lines.append(_hr())

    must_fix = [c for c in checklist_items if c["priority"] == "MUST FIX"]
    should_fix = [c for c in checklist_items if c["priority"] == "SHOULD FIX"]
    reminders = [c for c in checklist_items if c["priority"] == "REMINDER"]

    num = 1
    if must_fix:
        lines.append(_row("MUST FIX:"))
        for item in must_fix:
            text = _truncate(f"[{item['source_display']}] {item['item']}", _BOX_WIDTH - 12)
            lines.append(_row(f"  {num}. {text}"))
            num += 1

    if should_fix:
        lines.append(_row(""))
        lines.append(_row("SHOULD FIX:"))
        for item in should_fix[:10]:  # Limit console display
            text = _truncate(f"[{item['source_display']}] {item['item']}", _BOX_WIDTH - 12)
            lines.append(_row(f"  {num}. {text}"))
            num += 1
        if len(should_fix) > 10:
            lines.append(_row(f"  ... and {len(should_fix) - 10} more (see Excel)"))

    if reminders:
        lines.append(_row(""))
        lines.append(_row("REMINDERS:"))
        for item in reminders:
            text = _truncate(item["item"], _BOX_WIDTH - 10)
            lines.append(_row(f"  - {text}"))

    lines.append(_hr())
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Excel report
# ---------------------------------------------------------------------------

def _write_excel(
    unified: dict,
    readiness_result: dict,
    client_issues: list[dict],
    checklist_items: list[dict],
    missing_sources: list[str],
    output_path: Path,
    cc_rows: list[dict] | None = None,
    prov_rows: list[dict] | None = None,
) -> bool:
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("WARNING: openpyxl not installed — Excel report skipped.")
        return False

    wb = Workbook()

    # Sheet 1: Executive Summary
    ws1 = wb.active
    ws1.title = "Executive Summary"
    _write_executive_sheet(ws1, unified, readiness_result)

    # Sheet 2: Source Summary
    ws2 = wb.create_sheet("Source Summary")
    _write_source_summary_sheet(ws2, unified, readiness_result, missing_sources)

    # Sheet 3: Client Issue List
    ws3 = wb.create_sheet("Client Issue List")
    _write_issue_list_sheet(ws3, client_issues)

    # Sheet 4: Detailed Findings
    ws4 = wb.create_sheet("Detailed Findings")
    _write_detailed_findings_sheet(ws4, unified)

    # Sheet 5: Cross-Source Validation
    ws5 = wb.create_sheet("Cross-Source Validation")
    _write_cross_source_sheet(ws5, unified)

    # Sheet 6: Resubmission Checklist
    ws6 = wb.create_sheet("Resubmission Checklist")
    _write_checklist_sheet(ws6, checklist_items)

    # Sheet 7: Phase Run Metadata
    ws7 = wb.create_sheet("Phase Run Metadata")
    _write_metadata_sheet(ws7, unified, readiness_result)

    # Sheet 8: Cost Center Summary
    ws8 = wb.create_sheet("Cost Center Summary")
    _write_cost_center_sheet(ws8, cc_rows or [])

    # Sheet 9: Provider Summary
    ws9 = wb.create_sheet("Provider Summary")
    _write_provider_summary_sheet(ws9, prov_rows or [])

    # Auto-width all sheets
    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=0)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 60)

    try:
        wb.save(output_path)
        return True
    except PermissionError:
        print(
            f"\nERROR: Cannot write Excel report — the file is open in another program.\n"
            f"  Close '{output_path.name}' in Excel and re-run Phase 5 to regenerate it.\n"
            f"  (The JSON manifest will still be written.)"
        )
        return False


def _write_executive_sheet(ws, unified: dict, readiness: dict) -> None:
    ws.append(["PIVOT Test File Review — Executive Summary"])
    ws.append([])
    ws.append(["Client", unified.get("client", "")])
    ws.append(["Round", unified.get("round", "")])
    ws.append(["Test Month", unified.get("test_month", "")])
    bf = unified.get("billing_format", "unknown")
    ws.append(["Billing Format", bf.title() if bf else "Unknown"])
    ws.append(["Month Aligned", "Yes" if unified.get("month_aligned", True) else "No"])
    ws.append([])
    ws.append(["Overall Readiness", readiness.get("overall", "")])
    ws.append(["Reason", readiness.get("reason", "")])
    ws.append([])

    pm = unified.get("phase_metadata", {})
    ws.append(["Phase Run Dates"])
    ws.append(["Phase 1", pm.get("phase1_date", "")])
    ws.append(["Phase 2", pm.get("phase2_date", "")])
    ws.append(["Phase 3", pm.get("phase3_date", "")])
    ws.append(["Phase 4", pm.get("phase4_date", "")])
    ws.append(["Phase 5", date.today().isoformat()])


def _write_source_summary_sheet(ws, unified: dict, readiness: dict, missing: list[str]) -> None:
    headers = ["Source", "Files", "Row Count", "Date Column", "Date Range", "Status", "CRITICAL", "HIGH", "MEDIUM", "LOW", "INFO", "Total Issues", "Phase 2 Compatible"]
    ws.append(headers)

    per_source = readiness.get("per_source", {})
    display_order = ["billing", "scheduling", "payroll", "gl", "quality", "patient_satisfaction", "cross_source"]

    for group in display_order:
        if group in per_source:
            ps = per_source[group]
            sdata = unified.get("sources", {}).get(group, {})
            files = ", ".join(sdata.get("files", [])) if group != "cross_source" else ""
            row_count = sdata.get("row_count", "") if group != "cross_source" else ""
            dr = sdata.get("date_range", {}) if group != "cross_source" else {}
            date_col = dr.get("date_column", "") or ""
            date_range = f"{dr['min']} to {dr['max']}" if dr.get("min") else ""
            ws.append([
                ps.get("display_name", group.title()),
                files,
                row_count,
                date_col,
                date_range,
                ps.get("status", ""),
                ps.get("critical", 0),
                ps.get("high", 0),
                ps.get("medium", 0),
                ps.get("low", 0),
                ps.get("info", 0),
                ps.get("total", 0),
                sdata.get("phase2_compatible", "N/A") if group != "cross_source" else "N/A",
            ])
        elif group in missing:
            ws.append([
                group.replace("_", " ").title(),
                "", "", "", "", "NOT SUBMITTED",
                "", "", "", "", "", "", "",
            ])

    # Total row
    tc = readiness.get("total_counts", {})
    ws.append([
        "TOTAL", "", "", "", "", "",
        tc.get("CRITICAL", 0), tc.get("HIGH", 0), tc.get("MEDIUM", 0),
        tc.get("LOW", 0), tc.get("INFO", 0),
        sum(v for k, v in tc.items() if k != "INFO"), "",
    ])


def _write_issue_list_sheet(ws, client_issues: list[dict]) -> None:
    headers = ["#", "Severity", "Source", "Phase", "Check", "Field", "Description", "Affected Rows", "Priority"]
    ws.append(headers)
    for i, issue in enumerate(client_issues, 1):
        ws.append([
            i,
            issue.get("severity", ""),
            issue.get("source_display", ""),
            f"Phase {issue.get('phase', '')}",
            issue.get("check", ""),
            issue.get("field", ""),
            issue.get("description", ""),
            issue.get("affected_rows", ""),
            issue.get("priority", ""),
        ])


def _write_detailed_findings_sheet(ws, unified: dict) -> None:
    headers = ["#", "Phase", "Source", "File", "Check", "Raw Column", "Staging Column",
               "Severity", "Requirement Level", "Message", "Affected Rows", "Deduplicated"]
    ws.append(headers)

    num = 1
    # Source findings
    for group, sdata in unified.get("sources", {}).items():
        display = sdata.get("display_name", group.title())
        files = ", ".join(sdata.get("files", []))
        for issue in sdata.get("issues", []):
            ws.append([
                num,
                f"Phase {issue.get('phase', '')}",
                display,
                files,
                issue.get("check", ""),
                issue.get("raw_column", ""),
                issue.get("staging_column", ""),
                issue.get("severity", ""),
                issue.get("requirement_level", ""),
                issue.get("message", ""),
                issue.get("affected_rows", ""),
                "Yes" if issue.get("deduplicated") else "",
            ])
            num += 1

    # Cross-source findings
    for issue in unified.get("cross_source_issues", []):
        ws.append([
            num,
            "Phase 4",
            " <-> ".join(issue.get("sources_involved", [])),
            issue.get("files_compared", ""),
            issue.get("check", ""),
            "", "",
            issue.get("severity", ""),
            "",
            issue.get("message", ""),
            "",
            "Yes" if issue.get("deduplicated") else "",
        ])
        num += 1


def _write_cross_source_sheet(ws, unified: dict) -> None:
    headers = ["Check", "Description", "Files Compared", "Severity", "Status", "Message"]
    ws.append(headers)

    for issue in unified.get("cross_source_issues", []):
        if issue.get("deduplicated"):
            continue
        check = issue.get("check", "")
        label = _PHASE4_CHECK_LABELS.get(check, check)
        sev = issue.get("severity", "INFO")
        status = "PASS" if sev in ("INFO", "PASS") else ("CONDITIONAL" if sev == "HIGH" else ("FAIL" if sev == "CRITICAL" else "REVIEW"))
        ws.append([
            check,
            label,
            issue.get("files_compared", ""),
            sev,
            status,
            issue.get("message", ""),
        ])


def _write_checklist_sheet(ws, checklist_items: list[dict]) -> None:
    headers = ["#", "Priority", "Source", "Action Item", "Related Issues"]
    ws.append(headers)
    for i, item in enumerate(checklist_items, 1):
        ws.append([
            i,
            item.get("priority", ""),
            item.get("source_display", ""),
            item.get("item", ""),
            ", ".join(item.get("finding_ids", [])),
        ])


def _write_cost_center_sheet(ws, cc_rows: list[dict]) -> None:
    headers = [
        "CostCenterNumber", "CostCenterName",
        "WorkRvu", "Charges", "Payment", "Adjustments",
        "CompletedAppointments",
        "ProviderHours", "ProviderAmount",
        "SupportStaffHours", "SupportStaffAmount",
        "GL_ProviderPayAmount", "GL_SupportStaffAmount",
        "GL_MedicalPracticeCharges", "GL_MedicalPracticeAdjustments",
        "QualityRecordCount",
    ]
    ws.append(headers)
    for row in cc_rows:
        ws.append([row.get(h, 0 if h not in ("CostCenterNumber", "CostCenterName") else "") for h in headers])


def _write_provider_summary_sheet(ws, prov_rows: list[dict]) -> None:
    headers = [
        "ProviderNPI", "ProviderName",
        "WorkRvu", "Charges", "Payment", "Adjustments",
        "CompletedAppointments",
        "PayrollHours", "PayrollAmount",
        "QualityRecordCount",
    ]
    ws.append(headers)
    for row in prov_rows:
        ws.append([row.get(h, 0 if h not in ("ProviderNPI", "ProviderName") else "") for h in headers])


def _write_metadata_sheet(ws, unified: dict, readiness: dict) -> None:
    ws.append(["Phase", "Date Run", "Key Result"])

    pm = unified.get("phase_metadata", {})
    files_count = sum(len(s.get("files", [])) for s in unified.get("sources", {}).values())
    bf = unified.get("billing_format", "unknown")

    ws.append(["Phase 1", pm.get("phase1_date", ""),
               f"{files_count} files, test month {unified.get('test_month', '')}, billing format: {bf}"])

    # Phase 2 summary
    p2_compat_list = []
    for group, sdata in unified.get("sources", {}).items():
        c = sdata.get("phase2_compatible", "N/A")
        if c != "N/A":
            p2_compat_list.append(f"{sdata.get('display_name', group)}: {c}")
    ws.append(["Phase 2", pm.get("phase2_date", ""),
               f"Compatibility: {', '.join(p2_compat_list) if p2_compat_list else 'N/A'}"])

    # Phase 3 summary
    tc = readiness.get("total_counts", {})
    ws.append(["Phase 3", pm.get("phase3_date", ""),
               f"Data quality checks across all sources"])

    # Phase 4 summary
    cross_issues = unified.get("cross_source_issues", [])
    active = [i for i in cross_issues if not i.get("deduplicated")]
    ws.append(["Phase 4", pm.get("phase4_date", ""),
               f"{len(active)} cross-source findings"])

    ws.append(["Phase 5", date.today().isoformat(),
               f"Readiness: {readiness.get('overall', 'Unknown')}"])


# ---------------------------------------------------------------------------
# JSON manifest
# ---------------------------------------------------------------------------

def _write_json(
    unified: dict,
    readiness_result: dict,
    client_issues: list[dict],
    checklist_items: list[dict],
    output_path: Path,
) -> None:
    # Build cross-source summary from Phase 4 data
    cross_summary = {}
    for issue in unified.get("cross_source_issues", []):
        check = issue.get("check", "")
        if check not in cross_summary:
            sev = issue.get("severity", "INFO")
            cross_summary[check] = {
                "severity": sev,
                "status": "PASS" if sev in ("INFO", "PASS") else ("CONDITIONAL" if sev == "HIGH" else "FAIL"),
                "message": issue.get("message", ""),
            }

    manifest = {
        "client": unified.get("client", ""),
        "round": unified.get("round", ""),
        "date_run": date.today().isoformat(),
        "test_month": unified.get("test_month", ""),
        "billing_format": unified.get("billing_format", ""),
        "readiness": readiness_result,
        "client_issues": client_issues,
        "resubmission_checklist": checklist_items,
        "cross_source_summary": cross_summary,
        "phase_metadata": unified.get("phase_metadata", {}),
        "source_summary": {
            group: {
                "display_name": sdata.get("display_name", ""),
                "files": sdata.get("files", []),
                "row_count": sdata.get("row_count", 0),
                "date_range": sdata.get("date_range", {}),
                "severity_counts": sdata.get("severity_counts", {}),
                "phase2_compatible": sdata.get("phase2_compatible", "N/A"),
            }
            for group, sdata in unified.get("sources", {}).items()
        },
    }

    with open(output_path, "w", encoding="utf-8") as fh:
        json.dump(manifest, fh, indent=2, default=str)


# ---------------------------------------------------------------------------
# Main render entry point
# ---------------------------------------------------------------------------

def render(
    unified: dict,
    readiness_result: dict,
    client_issues: list[dict],
    checklist_items: list[dict],
    missing_sources: list[str],
    output_dir: Path,
    client: str,
    round_id: str,
    cc_rows: list[dict] | None = None,
    prov_rows: list[dict] | None = None,
) -> None:
    """Full Phase 5 render: console + Excel + JSON."""
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    # Console
    print()
    summary = _render_executive_summary(unified, readiness_result)
    _safe_print(summary)
    print()

    issues_box = _render_issue_list(client_issues)
    _safe_print(issues_box)
    print()

    checklist_box = _render_checklist(checklist_items)
    _safe_print(checklist_box)
    print()

    # Excel
    today_str = date.today().strftime("%Y%m%d")
    xlsx_path = output_dir / f"{client}_{round_id}_Phase5_{today_str}.xlsx"
    if _write_excel(unified, readiness_result, client_issues, checklist_items, missing_sources, xlsx_path,
                    cc_rows=cc_rows, prov_rows=prov_rows):
        print(f"Excel report: {xlsx_path}")

    # JSON
    json_path = output_dir / "phase5_findings.json"
    _write_json(unified, readiness_result, client_issues, checklist_items, json_path)
    print(f"JSON manifest: {json_path}")


def _safe_print(text: str) -> None:
    safe = text.encode("ascii", errors="replace").decode("ascii")
    print(safe)
