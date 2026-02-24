"""
phase4/report.py

Console display, Excel report, and JSON manifest for Phase 4.
ASCII-only output (no Unicode box-drawing characters), matching Phase 3 convention.
"""

from __future__ import annotations

import json
from datetime import date
from pathlib import Path
from typing import Any

import pandas as pd


_BOX_WIDTH = 69
_SEV_ORDER = {"CRITICAL": 0, "HIGH": 1, "MEDIUM": 2, "INFO": 3, "PASS": 4}
_CHECK_LABELS = {
    "C0": "Trans <-> Charges",
    "C1": "Billing <-> GL",
    "C2": "Billing <-> Payroll",
    "C3": "Billing <-> Scheduling",
    "C4": "Payroll <-> GL",
    "C5": "Scheduling <-> GL",
}


# ---------------------------------------------------------------------------
# Console formatting helpers (matching phase3/report.py style)
# ---------------------------------------------------------------------------

def _sev_icon(sev: str) -> str:
    return {"CRITICAL": "X", "HIGH": "!", "MEDIUM": "o", "INFO": ".", "PASS": "."}.get(sev, ".")


def _hr(char: str = "-") -> str:
    return "+" + char * (_BOX_WIDTH - 2) + "+"


def _row(text: str) -> str:
    return "| " + text.ljust(_BOX_WIDTH - 4) + " |"


def _kv_row(key: str, val: str, key_width: int = 22) -> str:
    k = key.ljust(key_width)
    v = str(val).ljust(_BOX_WIDTH - key_width - 7)
    return f"| {k}| {v}|"


def _truncate(s: str, max_len: int = _BOX_WIDTH - 12) -> str:
    return s[:max_len] + "..." if len(s) > max_len else s


def _pass_status(finding: dict) -> str:
    if finding.get("skipped"):
        return "SKIPPED"
    sev = finding.get("severity", "INFO")
    if sev == "CRITICAL":
        return "FAIL"
    if sev == "HIGH":
        return "CONDITIONAL"
    if sev in ("MEDIUM", "INFO", "PASS"):
        return "PASS"
    return "PASS"


# ---------------------------------------------------------------------------
# Console rendering
# ---------------------------------------------------------------------------

def _render_finding_lines(finding: dict) -> list[str]:
    """Render finding lines for one check block."""
    lines = []
    sev = finding.get("severity", "INFO")
    msg = _truncate(finding.get("message", ""))
    icon = _sev_icon(sev)
    lines.append(_row(f"  {icon} {sev}: {msg}"))
    return lines


def _render_sub_finding_lines(sub_findings, prefix: str = "    ") -> list[str]:
    """Render sub-check finding lines."""
    lines = []
    if isinstance(sub_findings, dict):
        sub_findings = [sub_findings]
    for f in sub_findings:
        sev = f.get("severity", "INFO")
        msg = _truncate(f.get("message", ""), _BOX_WIDTH - len(prefix) - 15)
        icon = _sev_icon(sev)
        check_id = f.get("check", "")
        lines.append(_row(f"{prefix}{icon} {sev} [{check_id}]: {msg}"))
    return lines


def render_check_box(check_id: str, finding: dict) -> str:
    """Render console box for one check."""
    lines = []
    lines.append(_hr())
    lines.append(_row("CROSS-SOURCE VALIDATION"))
    lines.append(_hr("-"))
    label = _CHECK_LABELS.get(check_id, check_id)
    lines.append(_kv_row("Check", f"{check_id}: {label}"))
    lines.append(_kv_row("Files Compared", finding.get("files_compared", "N/A")[:44]))
    lines.append(_hr("-"))

    if finding.get("skipped"):
        lines.append(_row(f"  . INFO: {finding.get('message', 'Skipped')}"))
    else:
        # C0 has sub_checks dict
        if "sub_checks" in finding and isinstance(finding["sub_checks"], dict):
            for sub_key, sub_val in finding["sub_checks"].items():
                if isinstance(sub_val, list):
                    lines.extend(_render_sub_finding_lines(sub_val, "    "))
                else:
                    lines.extend(_render_sub_finding_lines(sub_val, "    "))
        # C1, C5 have findings list
        elif "findings" in finding and isinstance(finding["findings"], list):
            for f in finding["findings"]:
                lines.extend(_render_finding_lines(f))
        # C3 has sub_checks with mixed list/dict
        elif "sub_checks" in finding:
            sc = finding["sub_checks"]
            for sub_key in ("C3a", "C3b", "C3c"):
                sub_val = sc.get(sub_key)
                if sub_val is None:
                    continue
                if isinstance(sub_val, list):
                    lines.extend(_render_sub_finding_lines(sub_val, "    "))
                else:
                    lines.extend(_render_sub_finding_lines(sub_val, "    "))
        else:
            # Single finding
            lines.extend(_render_finding_lines(finding))

    lines.append(_hr())
    return "\n".join(lines)


def render_summary_table(all_findings: dict[str, dict]) -> str:
    """Render the Phase 4 cross-source summary table."""
    lines = []
    lines.append(_hr())
    lines.append(_row("CROSS-SOURCE SUMMARY"))
    lines.append(_hr("-"))

    header = f"| {'Check':<27}| {'HIGH':>6}| {'MEDIUM':>8}| {'Pass?':<12}|"
    lines.append(header)
    lines.append(_hr("-"))

    for check_id in ("C0", "C1", "C2", "C3", "C4", "C5"):
        finding = all_findings.get(check_id, {})
        label = _CHECK_LABELS.get(check_id, check_id)
        short = f"{check_id}: {label}"[:26]

        if finding.get("skipped") or not finding:
            lines.append(f"| {short:<27}| {'  -':>6}| {'  -':>8}| {'SKIPPED':<12}|")
            continue

        # Count severities across all nested findings
        high = _count_severity(finding, "HIGH")
        medium = _count_severity(finding, "MEDIUM")
        status = _pass_status(finding)

        lines.append(f"| {short:<27}| {high:>6}| {medium:>8}| {status:<12}|")

    lines.append(_hr())
    return "\n".join(lines)


def _count_severity(finding: dict, target: str) -> int:
    """
    Count leaf-level occurrences of a severity in a (potentially nested) finding.

    If a finding has sub_checks or a findings list, count only from those
    child items (not the top-level aggregate). If it has neither, count
    the top-level severity directly.
    """
    if finding.get("skipped"):
        return 0

    count = 0
    has_children = bool(finding.get("sub_checks")) or bool(finding.get("findings"))

    if not has_children:
        # Leaf-level finding — count the top-level severity
        if finding.get("severity") == target:
            count += 1
        return count

    # Sub-checks (C0, C3)
    sub_checks = finding.get("sub_checks", {})
    if isinstance(sub_checks, dict):
        for sub_val in sub_checks.values():
            if isinstance(sub_val, list):
                for f in sub_val:
                    if f.get("severity") == target:
                        count += 1
            elif isinstance(sub_val, dict):
                if sub_val.get("severity") == target:
                    count += 1

    # Findings list (C1, C5)
    for f in finding.get("findings", []):
        if f.get("severity") == target:
            count += 1

    return count


# ---------------------------------------------------------------------------
# Excel report
# ---------------------------------------------------------------------------

def write_excel_report(
    all_findings: dict[str, dict],
    output_path: Path,
) -> bool:
    """Write the Phase 4 Excel report with 7 sheets."""
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("WARNING: openpyxl not installed — Excel report skipped.")
        return False

    wb = Workbook()

    # ── Sheet 1: Trans-Charges (C0) ──────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Trans-Charges"
    c0 = all_findings.get("C0", {})
    _write_c0_sheet(ws1, c0)

    # ── Sheet 2: Billing-GL (C1) ─────────────────────────────────────────────
    ws2 = wb.create_sheet("Billing-GL")
    c1 = all_findings.get("C1", {})
    _write_c1_sheet(ws2, c1)

    # ── Sheet 3: Billing-Payroll (C2) ────────────────────────────────────────
    ws3 = wb.create_sheet("Billing-Payroll")
    c2 = all_findings.get("C2", {})
    _write_c2_sheet(ws3, c2)

    # ── Sheet 4: Billing-Scheduling (C3) ─────────────────────────────────────
    ws4 = wb.create_sheet("Billing-Scheduling")
    c3 = all_findings.get("C3", {})
    _write_c3_sheet(ws4, c3)

    # ── Sheet 5: Payroll-GL (C4) ─────────────────────────────────────────────
    ws5 = wb.create_sheet("Payroll-GL")
    c4 = all_findings.get("C4", {})
    _write_c4_sheet(ws5, c4)

    # ── Sheet 6: Scheduling-GL (C5) ──────────────────────────────────────────
    ws6 = wb.create_sheet("Scheduling-GL")
    c5 = all_findings.get("C5", {})
    _write_c5_sheet(ws6, c5)

    # ── Sheet 7: Cross-Source Summary ─────────────────────────────────────────
    ws7 = wb.create_sheet("Cross-Source Summary")
    _write_summary_sheet(ws7, all_findings)

    # Auto-width all sheets
    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=0)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 60)

    try:
        wb.save(output_path)
        return True
    except PermissionError:
        print(
            f"\nERROR: Cannot write Excel report — the file is open in another program.\n"
            f"  Close '{output_path.name}' in Excel and re-run Phase 4 to regenerate it.\n"
            f"  (The JSON manifest will still be written.)"
        )
        return False


def _ws_append_header(ws, headers: list[str]) -> None:
    ws.append(headers)


def _write_c0_sheet(ws, c0: dict) -> None:
    if c0.get("skipped") or not c0:
        ws.append(["Status", "Skipped"])
        ws.append(["Message", c0.get("message", "No C0 data")])
        return

    sub = c0.get("sub_checks", {})
    c0a = sub.get("C0a", {})
    c0b = sub.get("C0b", {})

    # C0a table
    ws.append(["C0a: Charge ID Linkage"])
    ws.append(["Severity", c0a.get("severity", ""), "Message", c0a.get("message", "")])
    ws.append(["Charge ID Column", c0a.get("charge_id_column", ""), "Transaction ID Column", c0a.get("transaction_id_column", "")])
    ws.append(["Charge Distinct", c0a.get("charge_distinct", ""), "Transaction Distinct", c0a.get("transaction_distinct", "")])
    ws.append(["Match Count", c0a.get("match_count", ""), "Match %", c0a.get("match_pct", "")])
    ws.append([])
    ws.append(["Unmatched Transaction ID Sample"])
    for v in c0a.get("unmatched_transaction_sample", []):
        ws.append([v])

    ws.append([])
    ws.append(["C0b: Payment Balance Reasonableness"])
    ws.append(["Severity", c0b.get("severity", ""), "Message", c0b.get("message", "")])
    ws.append(["Total Charges", c0b.get("total_charges", ""), "Zero Balance Count", c0b.get("zero_balance_count", "")])
    ws.append(["Zero Balance %", c0b.get("zero_balance_pct", ""), "Outstanding Balance Rate %", c0b.get("outstanding_balance_rate", "")])
    ws.append(["Total Charge Amount", c0b.get("total_charge_amount", ""), "Outstanding Balance", c0b.get("outstanding_balance", "")])
    ws.append(["Avg Outstanding Balance", c0b.get("avg_outstanding_balance", "")])


def _write_c1_sheet(ws, c1: dict) -> None:
    if c1.get("skipped") or not c1:
        ws.append(["Status", "Skipped"])
        ws.append(["Message", c1.get("message", "No C1 data")])
        return

    headers = ["Billing Column", "Severity", "Match %", "Matched Values", "Unmatched Values",
               "Total Charge Amount", "Unmatched Charge Amount", "Unmatched Charge %", "Message"]
    ws.append(headers)
    for f in c1.get("findings", []):
        ws.append([
            f.get("billing_column", ""),
            f.get("severity", ""),
            f.get("match_pct", ""),
            f.get("matched_count", ""),
            f.get("unmatched_count", ""),
            f.get("total_charge_amount", ""),
            f.get("unmatched_charge_amount", ""),
            f.get("unmatched_charge_pct", ""),
            f.get("message", ""),
        ])

    # Unmatched sample detail
    ws.append([])
    ws.append(["Unmatched Value Detail (top 20 per column)"])
    ws.append(["Billing Column", "Value", "Row Count", "Dollar Amount"])
    for f in c1.get("findings", []):
        for item in f.get("unmatched_sample", []):
            ws.append([
                f.get("billing_column", ""),
                item.get("value", ""),
                item.get("row_count", ""),
                item.get("dollar_amount", ""),
            ])


def _write_c2_sheet(ws, c2: dict) -> None:
    if c2.get("skipped") or not c2:
        ws.append(["Status", "Skipped"])
        ws.append(["Message", c2.get("message", "No C2 data")])
        return

    ws.append(["C2: Billing <-> Payroll Provider Match"])
    ws.append(["Severity", c2.get("severity", ""), "Match Method", c2.get("match_method", "")])
    ws.append(["Billing Providers Distinct", c2.get("billing_provider_distinct", ""),
               "Payroll Providers Distinct", c2.get("payroll_provider_distinct", "")])
    ws.append(["Exact Match Count", c2.get("exact_match_count", ""), "Exact Match %", c2.get("exact_match_pct", "")])
    ws.append(["Message", c2.get("message", "")])
    ws.append([])

    # Top providers
    ws.append(["Top Providers by Charge Volume"])
    ws.append(["NPI", "Name", "Charge Amount", "Matched"])
    for p in c2.get("top_providers", []):
        ws.append([p.get("npi", ""), p.get("name", ""), p.get("charge_amount", ""), str(p.get("matched", ""))])

    # Name match candidates
    cands = c2.get("name_match_candidates", [])
    if cands:
        ws.append([])
        ws.append(["Name Match Candidates (unmatched billing providers with fuzzy payroll match)"])
        ws.append(["Billing Name", "Payroll Candidate", "Fuzzy Score"])
        for c in cands:
            ws.append([c.get("billing_name", ""), c.get("payroll_name_candidate", ""), c.get("score", "")])


def _write_c3_sheet(ws, c3: dict) -> None:
    if c3.get("skipped") or not c3:
        ws.append(["Status", "Skipped"])
        ws.append(["Message", c3.get("message", "No C3 data")])
        return

    sub = c3.get("sub_checks", {})

    # C3a
    ws.append(["C3a: Location/Department Cross-Reference"])
    ws.append(["Billing Column", "Severity", "Billing Distinct", "Overlap Count", "Billing Coverage %", "Message"])
    for f in sub.get("C3a", []):
        ws.append([
            f.get("billing_column", ""),
            f.get("severity", ""),
            f.get("billing_distinct", ""),
            f.get("overlap_count", ""),
            f.get("billing_coverage_pct", ""),
            f.get("message", ""),
        ])

    ws.append([])
    ws.append(["C3a Fuzzy Candidates"])
    ws.append(["Billing Column", "Billing Value", "Scheduling Candidate", "Score"])
    for f in sub.get("C3a", []):
        for c in f.get("fuzzy_candidates", []):
            ws.append([f.get("billing_column", ""), c.get("source_value", ""), c.get("match_candidate", ""), c.get("score", "")])

    ws.append([])

    # C3b
    c3b = sub.get("C3b", {})
    ws.append(["C3b: Provider NPI Cross-Reference"])
    ws.append(["Severity", c3b.get("severity", ""), "Message", c3b.get("message", "")])
    ws.append(["Billing Distinct", c3b.get("billing_distinct", ""), "Scheduling Distinct", c3b.get("scheduling_distinct", "")])
    ws.append(["Overlap Count", c3b.get("overlap_count", ""), "Billing Coverage %", c3b.get("billing_coverage_pct", ""), "Scheduling Coverage %", c3b.get("scheduling_coverage_pct", "")])

    ws.append([])

    # C3c
    c3c = sub.get("C3c", {})
    ws.append(["C3c: Patient ID Cross-Reference"])
    ws.append(["Severity", c3c.get("severity", ""), "Message", c3c.get("message", "")])
    ws.append(["Billing Distinct", c3c.get("billing_distinct", ""), "Scheduling Distinct", c3c.get("scheduling_distinct", "")])
    ws.append(["Overlap Count", c3c.get("overlap_count", ""), "Billing Coverage %", c3c.get("billing_coverage_pct", ""), "Scheduling Coverage %", c3c.get("scheduling_coverage_pct", "")])
    if c3c.get("format_note"):
        ws.append(["Format Note", c3c.get("format_note", "")])


def _write_c4_sheet(ws, c4: dict) -> None:
    if c4.get("skipped") or not c4:
        ws.append(["Status", "Skipped"])
        ws.append(["Message", c4.get("message", "No C4 data")])
        return

    ws.append(["C4: Payroll <-> GL Department to Cost Center"])
    ws.append(["Severity", c4.get("severity", ""), "Message", c4.get("message", "")])
    ws.append(["Dept Distinct", c4.get("dept_distinct", ""), "Matched Depts", c4.get("matched_dept_count", ""), "Unmatched Depts", c4.get("unmatched_dept_count", "")])
    ws.append(["Match %", c4.get("match_pct", "")])
    if c4.get("auto_extracted_offset"):
        ws.append(["Auto-Extracted Offset", str(c4.get("auto_extracted_offset", "")), "Log", c4.get("extraction_log", "")])

    ws.append([])
    ws.append(["Unmatched Departments"])
    ws.append(["Dept ID", "Dept Name", "Row Count"])
    for item in c4.get("unmatched_sample", []):
        ws.append([item.get("dept_id", ""), item.get("dept_name", ""), item.get("row_count", "")])


def _write_c5_sheet(ws, c5: dict) -> None:
    if c5.get("skipped") or not c5:
        ws.append(["Status", "Skipped"])
        ws.append(["Message", c5.get("message", "No C5 data")])
        return

    headers = ["Scheduling Column", "Severity", "Location Distinct", "Exact Match Count",
               "Fuzzy Candidates", "Unmatched Count", "Match %", "Message"]
    ws.append(headers)
    for f in c5.get("findings", []):
        ws.append([
            f.get("scheduling_column", ""),
            f.get("severity", ""),
            f.get("location_distinct", ""),
            f.get("exact_match_count", ""),
            f.get("fuzzy_candidate_count", ""),
            f.get("unmatched_count", ""),
            f.get("match_pct", ""),
            f.get("message", ""),
        ])

    ws.append([])
    ws.append(["Fuzzy Candidates"])
    ws.append(["Scheduling Column", "Scheduling Value", "GL Candidate", "Score"])
    for f in c5.get("findings", []):
        for c in f.get("fuzzy_candidates", []):
            ws.append([f.get("scheduling_column", ""), c.get("scheduling_value", ""), c.get("gl_candidate", ""), c.get("score", "")])


def _write_summary_sheet(ws, all_findings: dict[str, dict]) -> None:
    headers = ["Check", "Description", "HIGH Count", "MEDIUM Count", "Pass Status", "Files Compared", "Skipped"]
    ws.append(headers)

    for check_id in ("C0", "C1", "C2", "C3", "C4", "C5"):
        finding = all_findings.get(check_id, {})
        label = _CHECK_LABELS.get(check_id, check_id)
        skipped = finding.get("skipped", True)
        high = _count_severity(finding, "HIGH") if not skipped else 0
        medium = _count_severity(finding, "MEDIUM") if not skipped else 0
        status = _pass_status(finding) if finding else "SKIPPED"
        files = finding.get("files_compared", "N/A")
        ws.append([check_id, label, high, medium, status, files, str(skipped)])


# ---------------------------------------------------------------------------
# JSON manifest
# ---------------------------------------------------------------------------

def write_json_manifest(
    all_findings: dict[str, dict],
    client: str,
    round_id: str,
    output_path: Path,
) -> None:
    """Write phase4_findings.json."""
    checks_run = []
    checks_skipped = []
    has_high = False

    for check_id in ("C0", "C1", "C2", "C3", "C4", "C5"):
        finding = all_findings.get(check_id, {})
        if finding.get("skipped") or not finding:
            checks_skipped.append(check_id)
        else:
            checks_run.append(check_id)
            if _count_severity(finding, "HIGH") > 0:
                has_high = True

    manifest = {
        "client": client,
        "round": round_id,
        "date_run": date.today().isoformat(),
        "checks_run": checks_run,
        "checks_skipped": checks_skipped,
        "overall_pass": not has_high,
        "findings": {k: v for k, v in all_findings.items()},
    }

    with open(output_path, "w", encoding="utf-8") as fh:
        json.dump(manifest, fh, indent=2, default=str)


# ---------------------------------------------------------------------------
# Main render entry point
# ---------------------------------------------------------------------------

def render(
    all_findings: dict[str, dict],
    output_dir: Path,
    client: str,
    round_id: str,
) -> None:
    """
    Full Phase 4 render:
    - Print per-check console boxes
    - Print overall summary table
    - Write Excel report
    - Write JSON manifest
    """
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    # Console: per-check boxes
    for check_id in ("C0", "C1", "C2", "C3", "C4", "C5"):
        finding = all_findings.get(check_id)
        if finding is None:
            continue
        box = render_check_box(check_id, finding)
        safe_box = box.encode("ascii", errors="replace").decode("ascii")
        print(safe_box)
        print()

    # Summary table
    summary = render_summary_table(all_findings)
    safe_summary = summary.encode("ascii", errors="replace").decode("ascii")
    print(safe_summary)

    # Excel report
    today_str = date.today().strftime("%Y%m%d")
    xlsx_path = output_dir / f"{client}_{round_id}_Phase4_{today_str}.xlsx"
    if write_excel_report(all_findings, xlsx_path):
        print(f"\nExcel report: {xlsx_path}")

    # JSON manifest
    json_path = output_dir / "phase4_findings.json"
    write_json_manifest(all_findings, client, round_id, json_path)
    print(f"JSON manifest: {json_path}")
