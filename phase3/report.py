"""
phase3/report.py

Console display, Excel report, and JSON manifest for Phase 3.
ASCII-only output (no Unicode box-drawing characters).
"""

from __future__ import annotations

import json
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any

import pandas as pd

_BOX_WIDTH = 69
_SEV_ORDER = {"CRITICAL": 0, "HIGH": 1, "MEDIUM": 2, "LOW": 3, "INFO": 4}


def _sev_icon(sev: str) -> str:
    return {"CRITICAL": "X", "HIGH": "!", "MEDIUM": "o", "LOW": ".", "INFO": "."}.get(sev, ".")


def _sev_label(sev: str) -> str:
    return sev


def _hr(char: str = "-") -> str:
    return "+" + char * (_BOX_WIDTH - 2) + "+"


def _row(text: str, pad: bool = True) -> str:
    if pad:
        return "| " + text.ljust(_BOX_WIDTH - 4) + " |"
    return "| " + text + " |"


def _kv_row(key: str, val: str, key_width: int = 22) -> str:
    k = key.ljust(key_width)
    v = val.ljust(_BOX_WIDTH - key_width - 7)
    return f"| {k}| {v}|"


def _sev_counts(findings: list[dict]) -> dict[str, int]:
    counts = {"CRITICAL": 0, "HIGH": 0, "MEDIUM": 0, "LOW": 0, "INFO": 0}
    for f in findings:
        sev = f.get("severity", "INFO")
        if sev in counts:
            counts[sev] += 1
    return counts


def _truncate(s: str, max_len: int = _BOX_WIDTH - 10) -> str:
    return s[:max_len] + "..." if len(s) > max_len else s


def render_file_box(
    fname: str,
    source: str,
    row_count: int,
    universal_findings: list[dict],
    source_findings: list[dict],
) -> str:
    """Render the console box for a single file."""
    lines = []
    lines.append(_hr())
    lines.append(_row("DATA QUALITY REVIEW"))
    lines.append(_hr())
    lines.append(_kv_row("File Name", Path(fname).name[:45]))
    lines.append(_kv_row("Source", source[:45]))
    lines.append(_kv_row("Total Records", f"{row_count:,}"))
    lines.append(_hr())

    # Universal findings
    lines.append(_row("UNIVERSAL CHECKS"))
    u_sorted = sorted(universal_findings, key=lambda f: _SEV_ORDER.get(f.get("severity", "INFO"), 4))
    for finding in u_sorted:
        sev = finding.get("severity", "INFO")
        msg = _truncate(finding.get("message", ""))
        icon = _sev_icon(sev)
        count = finding.get("total_missing") or finding.get("duplicate_row_count") or finding.get("affected_row_count") or ""
        count_str = f" [{count:,}]" if isinstance(count, int) else ""
        lines.append(_row(f"  {icon} {sev}: {msg}{count_str}"))

    if not universal_findings:
        lines.append(_row("  . No issues found"))

    lines.append(_hr())

    # Source-specific findings
    src_label = source.replace("_", " ").upper()
    lines.append(_row(f"SOURCE-SPECIFIC CHECKS ({src_label})"))
    s_sorted = sorted(source_findings, key=lambda f: _SEV_ORDER.get(f.get("severity", "INFO"), 4))
    for finding in s_sorted:
        sev = finding.get("severity", "INFO")
        msg = _truncate(finding.get("message", ""))
        icon = _sev_icon(sev)
        lines.append(_row(f"  {icon} {sev}: {msg}"))

    if not source_findings:
        lines.append(_row("  . No issues found"))

    lines.append(_hr())

    # Issue summary
    all_findings = universal_findings + source_findings
    counts = _sev_counts(all_findings)
    summary = (
        f"  CRITICAL: {counts['CRITICAL']}  |  "
        f"HIGH: {counts['HIGH']}  |  "
        f"MEDIUM: {counts['MEDIUM']}  |  "
        f"LOW/INFO: {counts['LOW'] + counts['INFO']}"
    )
    lines.append(_row("ISSUE SUMMARY"))
    lines.append(_row(summary))
    lines.append(_hr())

    return "\n".join(lines)


def render_summary_table(all_file_results: dict[str, dict]) -> str:
    """Render the overall Phase 3 summary table."""
    lines = []
    lines.append(_hr())
    lines.append(_row("DATA QUALITY SUMMARY"))
    lines.append(_hr("-"))

    col_widths = [27, 6, 6, 8, 5, 7]
    header = (
        f"| {'File':<27}| {'CRIT':>6}| {'HIGH':>6}| {'MEDIUM':>8}| {'LOW':>5}| {'Total':>7}|"
    )
    lines.append(header)
    lines.append(_hr("-"))

    total_crit = total_high = total_med = total_low = total_all = 0

    for fname, result in all_file_results.items():
        u = result.get("universal_findings", [])
        s = result.get("source_specific_findings", [])
        counts = _sev_counts(u + s)
        short_name = Path(fname).name[:26]
        crit = counts["CRITICAL"]
        high = counts["HIGH"]
        med = counts["MEDIUM"]
        low = counts["LOW"] + counts["INFO"]
        total = crit + high + med + low
        total_crit += crit
        total_high += high
        total_med += med
        total_low += low
        total_all += total
        line = f"| {short_name:<27}| {crit:>6}| {high:>6}| {med:>8}| {low:>5}| {total:>7}|"
        lines.append(line)

    lines.append(_hr("-"))
    total_line = (
        f"| {'TOTAL':<27}| {total_crit:>6}| {total_high:>6}| {total_med:>8}| {total_low:>5}| {total_all:>7}|"
    )
    lines.append(total_line)
    lines.append(_hr())
    grand = (
        f"| Total Issues: {total_all}  "
        f"(CRITICAL: {total_crit}, HIGH: {total_high}, MEDIUM: {total_med}, "
        f"LOW/INFO: {total_low})"
    )
    lines.append(grand.ljust(_BOX_WIDTH - 2) + "|" if len(grand) < _BOX_WIDTH - 1 else grand)
    lines.append(_hr())
    return "\n".join(lines)


def _findings_to_rows(fname: str, findings: list[dict], check_type: str) -> list[dict]:
    rows = []
    for f in findings:
        rows.append({
            "File": Path(fname).name,
            "Check Type": check_type,
            "Check ID": f.get("check", ""),
            "Raw Column": f.get("raw_column", ""),
            "Staging Column": f.get("staging_column", ""),
            "Severity": f.get("severity", ""),
            "Message": f.get("message", ""),
            "Affected Count": (
                f.get("total_missing") or f.get("duplicate_row_count") or
                f.get("out_of_range_count") or f.get("affected_row_count") or ""
            ),
            "Sample Values / Notes": str(f.get("sample_rows") or f.get("sample_values") or "")[:200],
        })
    return rows


def write_excel_report(
    all_file_results: dict[str, dict],
    output_path: Path,
) -> None:
    """Write the Phase 3 Excel report with all 6 sheets."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("WARNING: openpyxl not installed — Excel report skipped.")
        return

    wb = Workbook()

    # ─── Sheet 1: Universal Findings ──────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Universal Findings"
    univ_rows = []
    for fname, result in all_file_results.items():
        univ_rows.extend(_findings_to_rows(fname, result.get("universal_findings", []), "Universal"))
    if univ_rows:
        df_u = pd.DataFrame(univ_rows)
        for r in [list(df_u.columns)] + df_u.values.tolist():
            ws1.append([str(v) if not isinstance(v, (int, float, type(None))) else v for v in r])

    # ─── Sheet 2: Source-Specific Findings ────────────────────────────────────
    ws2 = wb.create_sheet("Source-Specific Findings")
    src_rows = []
    for fname, result in all_file_results.items():
        src_rows.extend(_findings_to_rows(fname, result.get("source_specific_findings", []), "Source-Specific"))
    if src_rows:
        df_s = pd.DataFrame(src_rows)
        for r in [list(df_s.columns)] + df_s.values.tolist():
            ws2.append([str(v) if not isinstance(v, (int, float, type(None))) else v for v in r])

    # ─── Sheet 3: Null Analysis ────────────────────────────────────────────────
    ws3 = wb.create_sheet("Null Analysis")
    null_headers = ["File", "Raw Column", "Staging Column", "Requirement Level",
                    "Null Count", "Blank Count", "Total Missing", "Missing %",
                    "Charge Rows Only", "Severity", "Message"]
    ws3.append(null_headers)
    for fname, result in all_file_results.items():
        for f in result.get("universal_findings", []):
            if f.get("check") == "null_blank":
                ws3.append([
                    Path(fname).name,
                    f.get("raw_column", ""), f.get("staging_column", ""),
                    f.get("requirement_level", ""),
                    f.get("null_count", 0), f.get("blank_count", 0),
                    f.get("total_missing", 0), f.get("missing_pct", 0),
                    str(f.get("charge_rows_only", False)),
                    f.get("severity", ""), f.get("message", ""),
                ])

    # ─── Sheet 4: Duplicate Analysis ──────────────────────────────────────────
    ws4 = wb.create_sheet("Duplicate Analysis")
    dup_headers = ["File", "Check", "Key Columns", "Duplicate Row Count",
                   "Duplicate Group Count", "Sample Groups", "Severity", "Message"]
    ws4.append(dup_headers)
    for fname, result in all_file_results.items():
        for f in result.get("universal_findings", []):
            if f.get("check") in ("duplicate_records", "full_row_duplicates"):
                ws4.append([
                    Path(fname).name,
                    f.get("check", ""),
                    str(f.get("key_columns", "")),
                    f.get("duplicate_row_count", ""),
                    f.get("duplicate_group_count", ""),
                    str(f.get("sample_groups", ""))[:300],
                    f.get("severity", ""), f.get("message", ""),
                ])

    # ─── Sheet 5: Cost Center P&L ─────────────────────────────────────────────
    ws5 = wb.create_sheet("Cost Center P&L")
    pl_headers = [
        "File", "Cost Center #", "Cost Center Name",
        "Charges", "Adjustments", "Other Revenue", "Net Revenue",
        "Provider Comp", "Support Staff Comp", "Facilities", "Medical Supplies",
        "Other OpEx", "Total Expenses", "Net Income",
        "Required Present (of 5)", "Missing Required Categories", "Completeness Tier",
    ]
    ws5.append(pl_headers)
    for fname, result in all_file_results.items():
        for f in result.get("source_specific_findings", []):
            if f.get("check") == "G7" and "cost_center_pl" in f:
                for cc in f["cost_center_pl"]:
                    amts = cc.get("amounts", {})
                    ws5.append([
                        Path(fname).name,
                        cc.get("cost_center_number", ""),
                        cc.get("cost_center_name", ""),
                        amts.get("Charges", 0),
                        amts.get("Adjustments", 0),
                        amts.get("Other Revenue", 0),
                        amts.get("Net Revenue", 0),
                        amts.get("Provider Compensation", 0),
                        amts.get("Support Staff Compensation", 0),
                        amts.get("Facilities / Occupancy", 0),
                        amts.get("Medical Supplies", 0),
                        amts.get("Other Operating Expenses", 0),
                        amts.get("Total Expenses", 0),
                        amts.get("Net Income", 0),
                        cc.get("required_present", 0),
                        ", ".join(cc.get("required_missing", [])),
                        cc.get("completeness_tier", ""),
                    ])

    # ─── Sheet 6: Data Quality Summary ────────────────────────────────────────
    ws6 = wb.create_sheet("Data Quality Summary")
    summary_headers = ["File", "Source", "Record Count", "CRITICAL", "HIGH", "MEDIUM", "LOW", "INFO", "Total Issues"]
    ws6.append(summary_headers)
    for fname, result in all_file_results.items():
        u = result.get("universal_findings", [])
        s = result.get("source_specific_findings", [])
        counts = _sev_counts(u + s)
        ws6.append([
            Path(fname).name,
            result.get("source", ""),
            result.get("record_count", 0),
            counts["CRITICAL"], counts["HIGH"], counts["MEDIUM"],
            counts["LOW"], counts["INFO"],
            sum(counts.values()),
        ])

    # Auto-width columns in all sheets
    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=0)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 60)

    try:
        wb.save(output_path)
    except PermissionError:
        print(
            f"\nERROR: Cannot write Excel report — the file is open in another program.\n"
            f"  Close '{output_path.name}' in Excel and re-run Phase 3 to regenerate it.\n"
            f"  (The JSON manifest will still be written.)"
        )
        return False
    return True


def write_json_manifest(
    all_file_results: dict[str, dict],
    client: str,
    round_id: str,
    output_path: Path,
) -> None:
    """Write phase3_findings.json."""
    total_issues = 0
    total_critical = 0
    files_out = {}

    for fname, result in all_file_results.items():
        u = result.get("universal_findings", [])
        s = result.get("source_specific_findings", [])
        counts = _sev_counts(u + s)
        total_issues += sum(counts.values())
        total_critical += counts["CRITICAL"]

        files_out[fname] = {
            "source": result.get("source", ""),
            "record_count": result.get("record_count", 0),
            "universal_findings": u,
            "source_specific_findings": s,
            "cross_source_prep": result.get("cross_source_prep", {}),
            "issue_summary": {
                "critical": counts["CRITICAL"],
                "high": counts["HIGH"],
                "medium": counts["MEDIUM"],
                "low": counts["LOW"],
                "info": counts["INFO"],
            },
        }

    manifest = {
        "client": client,
        "round": round_id,
        "date_run": date.today().isoformat(),
        "overall_issue_count": total_issues,
        "overall_critical_count": total_critical,
        "files": files_out,
    }

    with open(output_path, "w", encoding="utf-8") as fh:
        json.dump(manifest, fh, indent=2, default=str)


def render(
    all_file_results: dict[str, dict],
    output_dir: Path,
    client: str,
    round_id: str,
) -> None:
    """
    Full Phase 3 render:
    - Print per-file console boxes
    - Print overall summary table
    - Write Excel report
    - Write JSON manifest
    """
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    # Console output
    for fname, result in all_file_results.items():
        box = render_file_box(
            fname=fname,
            source=result.get("source", "unknown"),
            row_count=result.get("record_count", 0),
            universal_findings=result.get("universal_findings", []),
            source_findings=result.get("source_specific_findings", []),
        )
        # Encode safely for Windows consoles
        safe_box = box.encode("ascii", errors="replace").decode("ascii")
        print(safe_box)
        print()

    # Summary table
    summary = render_summary_table(all_file_results)
    safe_summary = summary.encode("ascii", errors="replace").decode("ascii")
    print(safe_summary)

    # Excel report
    today_str = date.today().strftime("%Y%m%d")
    xlsx_path = output_dir / f"{client}_{round_id}_Phase3_{today_str}.xlsx"
    if write_excel_report(all_file_results, xlsx_path):
        print(f"\nExcel report: {xlsx_path}")

    # JSON manifest
    json_path = output_dir / "phase3_findings.json"
    write_json_manifest(all_file_results, client, round_id, json_path)
    print(f"JSON manifest: {json_path}")
